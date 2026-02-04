#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Substitui o template de planilhas .xlsx preservando o nome do arquivo.
Para arquivos da CLASSE II, renomeia para: RINEX_GPS_XXXX_DIA_DDD.xlsx

Uso (interativo):
    python renomear.py          # Dry-run (não grava)
    python renomear.py --commit # Grava por cima dos arquivos

Também aceita:
    python renomear.py "<pasta_raiz>" [--template "<novo_template.xlsx>"] [--sheet "Aba"] [--commit]
"""

import argparse
import os
import re
import shutil
import sys
import tempfile
from datetime import datetime, date
from typing import List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# === Template fixo (UNC) ===
DEFAULT_TEMPLATE = r"\\192.168.2.28\i\80225_PROJETO_IAT_PARANA\3 Execução de voo\GPS_0231_DIA_xxx.xlsx"

# ====== MAPEAMENTOS ANTIGO -> NOVO ======
MAPPINGS: List[Tuple[str, str]] = [
    ("J6", "J8"),
    ("J8", "I10"),
    ("K7", "K9"),
    ("B13:I13", "B15:I15"),  # linha 13 (B..H) do antigo -> linha 15 (B..H) no novo
]

# Células especiais no NOVO template:
CELL_DATE_IN_NEW = "J8"     # data base para o DDD
CELL_JULIAN_IN_NEW = "C11"  # receberá "DDD_2025"


def parse_args():
    ap = argparse.ArgumentParser(description="Substituir template de planilhas .xlsx")
    ap.add_argument("root", nargs="?", help="Pasta raiz para procurar .xlsx")
    ap.add_argument("--template", help="Override do caminho do template .xlsx", default=None)
    ap.add_argument("--sheet", help="Nome da aba a usar (se ausente, usa a ativa)", default=None)
    ap.add_argument("--commit", action="store_true", help="Grava alterações (sem este flag é dry-run)")
    args = ap.parse_args()

    # Se não informar root, entra no modo interativo só para escolher a pasta
    if not args.root:
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk(); root.withdraw()
            args.root = filedialog.askdirectory(title="Escolha a PASTA RAIZ")
            root.destroy()
            # 👉 Se o usuário escolheu pela interface, grava automaticamente
            args.commit = True
        except Exception:
            args.root = input("Pasta raiz (root): ").strip('"').strip()
            # Também grava automaticamente no modo texto interativo
            args.commit = True

    # Resolve o caminho do template: prioridade --template > DEFAULT_TEMPLATE
    args.new_template = args.template or DEFAULT_TEMPLATE

    # Validações
    if not args.root or not os.path.isdir(args.root):
        ap.error("Pasta raiz inválida ou não fornecida.")
    if not args.new_template or not os.path.isfile(args.new_template):
        ap.error(f"Arquivo de template inválido ou não encontrado: {args.new_template}")

    return args



def iter_xlsx(root_dir: str):
    for dirpath, _, filenames in os.walk(root_dir):
        for fn in filenames:
            if fn.lower().endswith(".xlsx") and not fn.startswith("~$"):
                yield os.path.join(dirpath, fn)


def load_ws(path: str, sheet_name: Optional[str]):
    wb = load_workbook(path, data_only=True)  # lê valores (não fórmulas)
    if sheet_name is None:
        ws = wb.active
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada em: {path}")
        ws = wb[sheet_name]
    return wb, ws


def merged_top_left(ws, row: int, col: int):
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = mr.bounds
        if (min_r <= row <= max_r) and (min_c <= col <= max_c):
            return (min_r, min_c)
    return None


def set_value_safe(ws, row: int, col: int, value):
    tl = merged_top_left(ws, row, col)
    if tl is not None:
        r, c = tl
        ws.cell(row=r, column=c).value = value
    else:
        ws.cell(row=row, column=col).value = value


def ensure_same_shape(src_range: str, dst_range: str):
    def shape(rng: str):
        min_c, min_r, max_c, max_r = range_boundaries(rng)
        return (max_r - min_r + 1, max_c - min_c + 1)
    s_rows, s_cols = shape(src_range)
    d_rows, d_cols = shape(dst_range)
    if (s_rows, s_cols) != (d_rows, d_cols):
        raise ValueError(
            f"Intervalos com formas diferentes: {src_range} ({s_rows}x{s_cols}) vs {dst_range} ({d_rows}x{d_cols})"
        )
    return s_rows, s_cols


def copy_range(src_ws, dst_ws, src_range: str, dst_range: str):
    s_rows, s_cols = ensure_same_shape(src_range, dst_range)
    s_min_c, s_min_r, _, _ = range_boundaries(src_range)
    d_min_c, d_min_r, _, _ = range_boundaries(dst_range)
    for ri in range(s_rows):
        for ci in range(s_cols):
            val = src_ws.cell(row=s_min_r + ri, column=s_min_c + ci).value
            set_value_safe(dst_ws, d_min_r + ri, d_min_c + ci, val)


def copy_cell(src_ws, dst_ws, src_addr: str, dst_addr: str):
    copy_range(src_ws, dst_ws, src_addr, dst_addr)


def compute_julian_str_from_cell(ws, date_cell_addr: str, year_fixed: int = 2025) -> str:
    val = ws[date_cell_addr].value
    if val is None:
        raise ValueError(f"Data ausente em {date_cell_addr}")

    if isinstance(val, datetime):
        dt = val.date()
    elif isinstance(val, date):
        dt = val
    elif isinstance(val, (int, float)):
        base = date(1899, 12, 30)  # base Excel (Windows)
        dt = date.fromordinal(base.toordinal() + int(val))
    elif isinstance(val, str):
        s = val.strip()
        dt = None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y", "%d-%m-%y", "%d.%m.%y"):
            try:
                dt = datetime.strptime(s, fmt).date()
                break
            except Exception:
                pass
        if dt is None:
            for fmt in ("%d/%m", "%d-%m", "%d.%m"):
                try:
                    base_d = datetime.strptime(s, fmt).date()
                    dt = date(year_fixed, base_d.month, base_d.day)
                    break
                except Exception:
                    pass
        if dt is None:
            raise ValueError(f"Não consegui interpretar a data '{val}' em {date_cell_addr}")
    else:
        raise ValueError(f"Tipo de data não suportado em {date_cell_addr}: {type(val)}")

    dt = date(year_fixed, dt.month, dt.day)  # força ano 2025
    ddd = dt.timetuple().tm_yday
    return f"{ddd:03d}_{year_fixed}"


def write_julian_safe(ws, julian_str: str, addr: str):
    min_c, min_r, max_c, max_r = range_boundaries(addr if ":" in addr else f"{addr}:{addr}")
    set_value_safe(ws, min_r, min_c, julian_str)


def is_class_ii(path: str) -> bool:
    # normaliza e verifica "2_CLASSE_II" no caminho
    p = path.replace("\\", "/").lower()
    return "/2_classe_ii/" in p


def parse_gps_from_filename(filename: str) -> Optional[str]:
    """
    tenta extrair XXXX de nomes do tipo GPS_XXXX_DIA_DDD(.xlsx)
    retorna somente os dígitos de XXXX se encontrar
    """
    name = os.path.splitext(os.path.basename(filename))[0]
    m = re.search(r"gps[_\- ]?(\d+)[_\- ]+dia[_\- ]?(\d+)", name, re.IGNORECASE)
    if m:
        gps_num = m.group(1)
        return gps_num
    # fallback: só GPS_XXXX
    m2 = re.search(r"gps[_\- ]?(\d+)", name, re.IGNORECASE)
    if m2:
        return m2.group(1)
    return None


def ensure_three_digits(ddd: str) -> str:
    # normaliza DDD para 3 dígitos
    d = int(ddd)
    return f"{d:03d}"


def ask_user(prompt: str) -> str:
    # pergunta no terminal (fallback simples)
    return input(prompt).strip()


def maybe_rename_class_ii(old_path: str, julian_str: str, commit: bool):
    """
    Se for CLASSE II, renomeia o arquivo para RINEX_GPS_XXXX_DIA_DDD.xlsx
    DDD vindo de julian_str (formato 'DDD_2025').
    Tenta extrair XXXX do nome atual; se não conseguir, pergunta ao usuário.
    """
    if not is_class_ii(old_path):
        return  # só renomeia na Classe II

    dirname = os.path.dirname(old_path)
    base_no_ext = os.path.splitext(os.path.basename(old_path))[0]

    # extrai DDD do julian_str (ex.: '223_2025' -> '223')
    ddd = julian_str.split("_", 1)[0]
    ddd = ensure_three_digits(ddd)

    gps_num = parse_gps_from_filename(base_no_ext)
    if not gps_num:
        # pergunta ao usuário
        gps_num = ask_user(f"Não consegui identificar o número do GPS para '{base_no_ext}'. Informe apenas os dígitos (ex.: 6319): ").strip()
        gps_num = re.sub(r"\D+", "", gps_num)
        if not gps_num:
            print(f"[AVISO] GPS não informado/identificado. Mantendo nome original.")
            return

    new_name = f"RINEX_GPS_{gps_num}_DIA_{ddd}.xlsx"
    new_path = os.path.join(dirname, new_name)

    if os.path.abspath(new_path) == os.path.abspath(old_path):
        # já está com o nome certo
        return

    if commit:
        try:
            os.replace(old_path, new_path)  # substitui se já existir
            print(f"[RENOMEADO] {os.path.basename(old_path)} -> {new_name}")
        except Exception as e:
            print(f"[ERRO-RENOMEAR] {old_path}: {e}", file=sys.stderr)
    else:
        print(f"[DRY-RUN] Renomearia: {os.path.basename(old_path)} -> {new_name}")


def process_file(old_path: str, template_path: str, sheet_name: Optional[str], commit: bool):
    print(f"\n==> Processando: {old_path}")

    old_wb, old_ws = load_ws(old_path, sheet_name)
    new_wb, new_ws = load_ws(template_path, sheet_name)

    # Copia conforme MAPPINGS (merge-safe)
    for src, dst in MAPPINGS:
        copy_range(old_ws, new_ws, src, dst)
        print(f"[MAP] {src} -> {dst}")

    # C11 = DDD_2025 a partir da data em J8 (no NOVO)
    julian_str = compute_julian_str_from_cell(new_ws, CELL_DATE_IN_NEW, year_fixed=2025)
    write_julian_safe(new_ws, julian_str, CELL_JULIAN_IN_NEW)
    print(f"[SET] {CELL_JULIAN_IN_NEW} <- {julian_str} (com base em {CELL_DATE_IN_NEW})")

    # Salvar por cima do antigo
    if commit:
        dirname = os.path.dirname(old_path)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=dirname) as tmp:
            tmp_path = tmp.name
        new_wb.save(tmp_path)
        shutil.move(tmp_path, old_path)
        print(f"[OK] Substituído: {old_path}")
    else:
        print("[DRY-RUN] (não gravado)")

    # Fechar livros
    old_wb.close()
    new_wb.close()

    # Renomear se for CLASSE II (usa o arquivo já substituído em old_path)
    maybe_rename_class_ii(old_path, julian_str, commit)


def main():
    args = parse_args()

    total = 0
    for xlsx in iter_xlsx(args.root):
        try:
            process_file(xlsx, args.new_template, args.sheet, args.commit)
            total += 1
        except Exception as e:
            print(f"[ERRO] {xlsx}: {e}", file=sys.stderr)

    print(f"\nConcluído. Arquivos processados: {total}. Modo: {'COMMIT' if args.commit else 'DRY-RUN'}")


if __name__ == "__main__":
    main()
