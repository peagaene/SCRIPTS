import os
import pandas as pd
import geopandas as gpd
from openpyxl import load_workbook
from shutil import copyfile

# === CONFIGURAÇÕES ===
GPKG_PATH = r"D:\80225_PROJETO_IAT_PARANA\2 Planejamento voo\02 - VOO LASER\LOTE_10\1_1_GPKG\ES_PV_LASER_L10_FAIXAS_R0.gpkg"
PLANILHA_ORIGINAL = r"D:\80225_PROJETO_IAT_PARANA\2 Planejamento voo\02 - VOO LASER\APOIO\LOTE_10\ES_PV_LASER_L10_R0.xlsx"
PLANILHA_DESTINO = r"D:\80225_PROJETO_IAT_PARANA\2 Planejamento voo\02 - VOO LASER\LOTE_10\1_3_XLSX\\ES_PV_LASER_L10_R0.xlsx"

# === LEITURA DO GPKG ===
gdf = gpd.read_file(GPKG_PATH)

# Dividir coordenadas em LAT/LON separadas para os dois pontos
def dividir_coordenadas(dms_str):
    partes = dms_str.split()
    return partes[0], partes[1] if len(partes) == 2 else ("", "")

# Aplicar conversão
dados_faixas = gdf[[
    "FlightLine", "Length(m)", "AltMSL(m)", "MinAltAGL(m)", "MaxAltAGL(m)",
    "FirstEvent", "LastEventW", "Estimateddf"
]].copy()

dados_faixas[["FirstEvent_LAT", "FirstEvent_LON"]] = dados_faixas["FirstEvent"].apply(lambda x: pd.Series(dividir_coordenadas(x)))
dados_faixas[["LastEvent_LAT", "LastEvent_LON"]] = dados_faixas["LastEventW"].apply(lambda x: pd.Series(dividir_coordenadas(x)))

# Converter comprimento para km
dados_faixas["Length(m)"] = dados_faixas["Length(m)"] / 1000

# Reorganizar colunas
dados_final = dados_faixas[[
    "FlightLine", "Length(m)", "AltMSL(m)", "MinAltAGL(m)", "MaxAltAGL(m)",
    "FirstEvent_LAT", "FirstEvent_LON",
    "LastEvent_LAT", "LastEvent_LON",
    "Estimateddf"
]]

# === COPIA DA PLANILHA ORIGINAL ===
copyfile(PLANILHA_ORIGINAL, PLANILHA_DESTINO)

# === INSERÇÃO DOS DADOS PRESERVANDO FORMATO ===
wb = load_workbook(PLANILHA_DESTINO)
ws = wb.active

start_row = 9
for idx, row in dados_final.iterrows():
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=start_row + idx, column=col_idx, value=value)

wb.save(PLANILHA_DESTINO)
print("Nova planilha salva em:", PLANILHA_DESTINO)
