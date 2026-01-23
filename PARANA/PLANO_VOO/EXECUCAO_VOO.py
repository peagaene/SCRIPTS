import os
import geopandas as gpd
import pandas as pd
from shapely.geometry import LineString
from datetime import datetime, timedelta
from openpyxl import load_workbook

# === CONFIGURAÇÕES ===
DIR_VOOS = r"D:\80225_PROJETO_IAT_PARANA\2 - Execução de voo\BLOCO_I"
CAMINHO_PLANO_VOO = r"D:\80225_PROJETO_IAT_PARANA\2 - Execução de voo\PLANO_DE_VOO\ES_PV_LASER_L09_FAIXAS_R0.shp"
DIR_SAIDA = r"D:\80225_PROJETO_IAT_PARANA\2 - Execução de voo\BLOCO_I"
CAMINHO_TEMPLATE = r"D:\80225_PROJETO_IAT_PARANA\2 - Execução de voo\ES_RB_LASER_LXX_R0_ddmmaaaa.xlsx"
EPSG_UTM22S = "EPSG:31982"
ALTURA_VOO = 1400

# Função para converter tempo GPS para datetime
def gps_time_to_datetime(week, seconds):
    gps_epoch = datetime(1980, 1, 6)
    return gps_epoch + timedelta(weeks=int(week), seconds=float(seconds))

# === PROCESSAMENTO ===
plano_voo = gpd.read_file(CAMINHO_PLANO_VOO).to_crs(EPSG_UTM22S)

for pasta in os.listdir(DIR_VOOS):
    dir_dia = os.path.join(DIR_VOOS, pasta)
    if not os.path.isdir(dir_dia):
        continue

    shp_voo = [f for f in os.listdir(dir_dia) if f.endswith(".shp")]
    if not shp_voo:
        continue

    gdf_faixas = gpd.read_file(os.path.join(dir_dia, shp_voo[0])).to_crs(EPSG_UTM22S)

    print(f"\n[INFO] Arquivo lido: {shp_voo[0]}")
    print("[INFO] Colunas disponíveis no shapefile:", gdf_faixas.columns.tolist())

    linhas_centrais = []
    dados_planilha = []
    blocos_encontrados = set()
    altitudes_m = []

    for idx, faixa in gdf_faixas.iterrows():
        centroide = faixa.geometry.centroid
        linha_mais_proxima = plano_voo.distance(centroide).idxmin()
        linha_voo = plano_voo.loc[linha_mais_proxima]

        linha_central = faixa.geometry.representative_point()
        linha_simulada = LineString([linha_central.buffer(0.01).bounds[:2], linha_central.buffer(0.01).bounds[2:]])

        bloco = linha_voo['BLOCO']
        blocos_encontrados.add(bloco)
        nome = linha_voo['FlightLine']

        gps_week = faixa.get('GPS_week', None)
        gps_start = faixa.get('GPS_Start_', None)
        gps_stop = faixa.get('GPS_Stop_T', None)

        if pd.notna(gps_week) and pd.notna(gps_start):
            hora_inicio_dt = gps_time_to_datetime(gps_week, gps_start)
            hora_inicio = hora_inicio_dt.strftime('%Y-%m-%dT%H:%M:%S')
        else:
            hora_inicio = ''

        if pd.notna(gps_week) and pd.notna(gps_stop):
            hora_fim_dt = gps_time_to_datetime(gps_week, gps_stop)
            hora_fim = hora_fim_dt.strftime('%Y-%m-%dT%H:%M:%S')
        else:
            hora_fim = ''

        altitude_m = faixa.get('Altitude_[', '')
        velocidade_nos = faixa.get('Speed_[m/s', '')

        if pd.notna(altitude_m):
            altitudes_m.append(altitude_m - ALTURA_VOO)
            altitude_pes = round(altitude_m * 3.28084, 1)
        else:
            altitude_pes = ''

        if pd.notna(velocidade_nos):
            velocidade_nos = round(velocidade_nos * 1.94384, 1)
        else:
            velocidade_nos = ''

        print(f"[DEBUG] {nome} | Início: {hora_inicio} | Fim: {hora_fim} | Altitude (ft): {altitude_pes} | Vel (nós): {velocidade_nos}")

        dados_planilha.append([
            nome, hora_inicio, hora_fim, altitude_pes, velocidade_nos, 20, 80, 250
        ])

        linhas_centrais.append({"geometry": linha_voo.geometry, "FlightLine": nome})

    dados_planilha.sort(key=lambda x: x[0])

    nome_dia = pasta.replace("_", "-")
    gdf_poligonos = gdf_faixas.copy()
    gdf_poligonos['FlightLine'] = [d[0] for d in dados_planilha]
    gdf_poligonos.to_file(os.path.join(DIR_SAIDA, f"{nome_dia}_FAIXAS.shp"))

    gdf_linhas = gpd.GeoDataFrame(linhas_centrais, crs=EPSG_UTM22S)
    gdf_linhas.to_file(os.path.join(DIR_SAIDA, f"{nome_dia}_LINHAS.shp"))

    if len(blocos_encontrados) == 1:
        bloco_input = list(blocos_encontrados)[0]
    else:
        print(f"[Atenção] Vários blocos encontrados: {sorted(blocos_encontrados)}")
        bloco_input = sorted(blocos_encontrados)[0]

    data_dt = datetime.strptime(pasta, "%d_%m_%y")
    data_formatada = data_dt.strftime("%d%m%Y")
    nome_arquivo_excel = f"ES_RB_LASER_L09_R0_{data_formatada}.xlsx"
    caminho_saida_excel = os.path.join(DIR_SAIDA, nome_arquivo_excel)

    alt_media_m = round(sum(altitudes_m) / len(altitudes_m)) if altitudes_m else 760
    dia_juliano = data_dt.timetuple().tm_yday

    wb = load_workbook(CAMINHO_TEMPLATE)
    ws = wb.active

    for i, linha in enumerate(dados_planilha):
        ws[f"B{17 + i}"] = linha[0]
        ws[f"C{17 + i}"] = linha[1]
        ws[f"D{17 + i}"] = linha[2]
        ws[f"E{17 + i}"] = linha[3]
        ws[f"F{17 + i}"] = linha[4]
        ws[f"G{17 + i}"] = linha[5]
        ws[f"H{17 + i}"] = linha[6]
        ws[f"I{17 + i}"] = linha[7]

    ws["F15"] = f"ALT. MÉDIA DA REGIÃO: {alt_media_m}"
    ws["I12"] = f"DIA JULIANO: {dia_juliano}"
    ws["I15"] = f"DATA: {data_dt.strftime('%d/%m/%Y')}"
    ws["G13"] = f"NUM. FAIXAS: {len(dados_planilha)}"

    wb.save(caminho_saida_excel)
    print(f"[OK] Processado {pasta} → {nome_arquivo_excel}")
