import os
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook

PADROES_BASE = {
    "154": ["B154", "BASE YYY", "BYYY", "BYYY - BASE", "Base yyy"],
    "153": ["B153", "BASE ZZZ", "BZZZ"]
}

DIR_RINEX_BASE = r"F:\80225_PROJETO_IAT_PARANA\4 Ponto de apoio\RINEX"
DIR_TEMPLATE = r"F:\80225_PROJETO_IAT_PARANA\3 Execução de voo"
DIR_ENTREGA = r"F:\80225_PROJETO_IAT_PARANA\3 Execução de voo\ENTREGA\2_APOIO_DE_CAMPO"

MUNICIPIOS = {"153": "Clevelândia", "154": "Chopinzinho"}
BASES = {"153": "BV-153", "154": "BV-154"}

def ajustar_hora(hora_str):
    h, m, s = map(int, hora_str.split(":"))
    nova = datetime(2000, 1, 1, h, m, s) - timedelta(hours=3)
    return nova.strftime("%H:%M:%S")

def log(msg):
    print(msg)

print("[INFO] Iniciando geração de planilhas de campo...")

for data_folder in sorted(os.listdir(DIR_RINEX_BASE)):
    print(f"\n📁 Pasta: {data_folder}")
    if not data_folder.startswith("2025"):
        continue

    data_dt = datetime.strptime(data_folder, "%Y%m%d")
    juliano = data_dt.timetuple().tm_yday
    ano = data_dt.year
    dir_data = os.path.join(DIR_RINEX_BASE, data_folder)

    for subpasta in os.listdir(dir_data):
        subpasta_upper = subpasta.upper()
        if "BASE" not in subpasta_upper:
            continue

        base_id = None
        for bid, padroes in PADROES_BASE.items():
            if any(p.upper() in subpasta_upper for p in padroes):
                base_id = bid
                break

        if base_id is None:
            log(f"[AVISO] Base não reconhecida: {subpasta}")
            continue

        base_nome = BASES.get(base_id, "BV-XXX")
        municipio = MUNICIPIOS.get(base_id, "Munic. Desconhecido")
        rinex_dir = os.path.join(dir_data, subpasta, "Rinex")
        if not os.path.isdir(rinex_dir):
            continue

        arquivos = os.listdir(rinex_dir)
        arquivo_obs = next((f for f in arquivos if f.lower().endswith((".o", ".25o"))), None)
        if not arquivo_obs:
            log(f"[AVISO] Nenhum arquivo .o ou .25o encontrado em {rinex_dir}")
            continue

        print(f"📄 Arquivo RINEX: {arquivo_obs}")
        caminho_obs = os.path.join(rinex_dir, arquivo_obs)
        with open(caminho_obs, "r") as f:
            linhas = f.readlines()

        hora_inicio = hora_fim = None
        for linha in linhas:
            if "TIME OF FIRST OBS" in linha:
                p = linha.split()
                hora_inicio = f"{int(p[3]):02d}:{int(p[4]):02d}:{int(float(p[5])):02d}"
            elif "TIME OF LAST OBS" in linha:
                p = linha.split()
                hora_fim = f"{int(p[3]):02d}:{int(p[4]):02d}:{int(float(p[5])):02d}"

        if not hora_inicio or not hora_fim:
            log("[AVISO] Horários não encontrados. Pulando.")
            continue

        hora_inicio = ajustar_hora(hora_inicio)
        hora_fim = ajustar_hora(hora_fim)

        nome_template = "GPS_0231_DIA_xxx.xlsx"
        caminho_template = os.path.join(DIR_TEMPLATE, nome_template)
        if not os.path.exists(caminho_template):
            log(f"[ERRO] Template não encontrado: {nome_template}")
            continue

        wb = load_workbook(caminho_template)
        ws = wb.active
        ws["J6"] = data_dt.strftime("%d/%m/%Y")
        ws["K7"] = municipio
        ws["B13"] = base_nome
        ws["CD9"] = juliano
        ws["D13"] = altura = "1,539 m"
        ws["E13"] = altura
        ws["G13"] = hora_inicio
        ws["H13"] = hora_fim

        # ✅ Pasta agora com nome apenas DIA_xxx_yyyy
        pasta_saida = os.path.join(DIR_ENTREGA, f"DIA_{juliano}_{ano}")
        os.makedirs(pasta_saida, exist_ok=True)

        nome_saida = f"GPS_0231_DIA_{juliano}.xlsx"
        wb.save(os.path.join(pasta_saida, nome_saida))
        print(f"✅ Planilha salva: {nome_saida}")

        for f in arquivos:
            origem = os.path.join(rinex_dir, f)
            ext = os.path.splitext(f)[1]
            destino = os.path.join(pasta_saida, f"0231_DIA_{juliano}{ext}")
            shutil.copy2(origem, destino)
            print(f"📥 Copiado: {f} → 0231_DIA_{juliano}{ext}")

        log(f"[OK] Concluído para {base_nome} - Dia {juliano}")

print("\n✅ [FIM] Todas as planilhas foram processadas.")
