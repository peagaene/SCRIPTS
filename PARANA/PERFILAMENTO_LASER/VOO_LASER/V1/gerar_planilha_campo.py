import os
import re
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook


def gerar_planilha_campo(
    dir_rinex_base=r"\\192.168.2.28\\i\\80225_PROJETO_IAT_PARANA\\4 Ponto de apoio\\RINEX",
    dir_template=r"\\192.168.2.28\i\80225_PROJETO_IAT_PARANA\3 Execução de voo",
    dir_saida=r"\\192.168.2.28\i\80225_PROJETO_IAT_PARANA\3 Execução de voo\ENTREGA\2_APOIO_DE_CAMPO",
    datas_filtradas=None,
    log_func=print
):
    processados = []
    datas_filtradas = set(datas_filtradas) if datas_filtradas else None
    MUNICIPIOS = {
        "153": "Clevelândia",
        "154": "Chopinzinho",
        "155": "Laranjeiras do Sul",
    }

    BASES = {
        "153": "BV-153",
        "154": "BV-154",
        "155": "BV-155",
    }

    def encontrar_pasta_rinex(caminho_data):
        for raiz, dirs, _ in os.walk(caminho_data):
            for dir_nome in dirs:
                if dir_nome.lower() == "rinex":
                    return os.path.join(raiz, dir_nome)
        return None

    def ajustar_hora(hora_str):
        h, m, s = map(int, hora_str.split(":"))
        nova = datetime(2000, 1, 1, h, m, s) - timedelta(hours=3)
        return nova.strftime("%H:%M:%S")

    def obter_base_id(nome_base):
        match = re.search(r"(\d{3})", nome_base)
        return match.group(1) if match else None

    for base_folder in sorted(os.listdir(dir_rinex_base)):
        caminho_base = os.path.join(dir_rinex_base, base_folder)
        if not os.path.isdir(caminho_base):
            continue

        base_id = obter_base_id(base_folder)
        if base_id not in BASES:
            log_func(f"[AVISO] Base não reconhecida: {base_folder}")
            continue

        base_nome = BASES[base_id]
        municipio = MUNICIPIOS.get(base_id, "Munic. Desconhecido")

        for data_folder in sorted(os.listdir(caminho_base)):
            if not data_folder.startswith("2025"):
                continue

            caminho_data = os.path.join(caminho_base, data_folder)
            if not os.path.isdir(caminho_data):
                continue

            try:
                data_dt = datetime.strptime(data_folder, "%Y%m%d")
            except ValueError:
                log_func(f"[AVISO] Nome de pasta de dia inválido em {base_nome}: {data_folder}")
                continue

            if datas_filtradas and data_dt.date() not in datas_filtradas:
                continue

            rinex_dir = encontrar_pasta_rinex(caminho_data)

            if not rinex_dir:
                log_func(f"[AVISO] Pasta 'Rinex' não encontrada em {caminho_data} (procura em subpastas)")
                continue

            arquivos = os.listdir(rinex_dir)
            arquivo_obs = next((f for f in arquivos if f.lower().endswith((".o", ".25o"))), None)
            if not arquivo_obs:
                log_func(f"[AVISO] Nenhum arquivo .o ou .25o encontrado em {rinex_dir}")
                continue

            caminho_obs = os.path.join(rinex_dir, arquivo_obs)
            with open(caminho_obs, "r") as f:
                linhas = f.readlines()

            hora_inicio = hora_fim = None
            for linha in linhas:
                if "TIME OF FIRST OBS" in linha:
                    p = linha.split()
                    hora_inicio = f"{int(p[3]):02d}:{int(p[4]):02d}:{int(float(p[5])):02d}"
                elif "TIME OF LAST OBS" in linha:
                    p = linha.split()
                    hora_fim = f"{int(p[3]):02d}:{int(p[4]):02d}:{int(float(p[5])):02d}"

            if not hora_inicio or not hora_fim:
                log_func("[AVISO] Horários não encontrados. Pulando.")
                continue

            hora_inicio = ajustar_hora(hora_inicio)
            hora_fim = ajustar_hora(hora_fim)

            juliano = data_dt.timetuple().tm_yday
            ano = data_dt.year

            nome_template = "GPS_0231_DIA_xxx.xlsx"
            caminho_template = os.path.join(dir_template, nome_template)
            if not os.path.exists(caminho_template):
                log_func(f"[ERRO] Template não encontrado: {nome_template}")
                continue

            wb = load_workbook(caminho_template)
            ws = wb.active
            ws["J6"] = data_dt.strftime("%d/%m/%Y")
            ws["K7"] = municipio
            ws["B13"] = base_nome
            ws["CD9"] = juliano
            altura = "1,539 m"
            ws["D13"] = altura
            ws["E13"] = altura
            ws["G13"] = hora_inicio
            ws["H13"] = hora_fim

            pasta_saida = os.path.join(dir_saida, f"DIA_{juliano}_{ano}")
            os.makedirs(pasta_saida, exist_ok=True)

            nome_saida = f"GPS_0231_DIA_{juliano}.xlsx"
            wb.save(os.path.join(pasta_saida, nome_saida))

            for f in arquivos:
                origem = os.path.join(rinex_dir, f)
                ext = os.path.splitext(f)[1]
                destino = os.path.join(pasta_saida, f"0231_DIA_{juliano}{ext}")
                shutil.copy2(origem, destino)

            log_func(f"[OK] Planilha de campo gerada para {base_nome} em {data_dt.strftime('%d/%m/%Y')}.")
            processados.append((data_dt.date(), base_nome, pasta_saida))

    return processados
