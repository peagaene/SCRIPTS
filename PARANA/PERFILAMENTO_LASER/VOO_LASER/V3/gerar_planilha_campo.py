
import unicodedata

def _strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def _extrair_altura_do_txt(conteudo: str):
    # Prefer lines that mention both 'altura' and 'antena'
    linhas = conteudo.splitlines()
    for linha in linhas:
        lnorm = _strip_accents(linha).lower().replace('  ', ' ')
        if 'altura' in lnorm and ('antena' in lnorm or 'antea' in lnorm):
            m = re.search(r'(\d{1,2}(?:[\.,]\d{1,3})?)\s*(?:m\b|metros?)?', linha)
            if m:
                val = m.group(1).replace(',', '.')
                try:
                    return round(float(val), 3)
                except:
                    pass
    # Fallback: first decimal number anywhere
    m = re.search(r'(\d{1,2}(?:[\.,]\d{1,3})?)', conteudo)
    if m:
        val = m.group(1).replace(',', '.')
        try:
            return round(float(val), 3)
        except:
            pass
    return None


def _procurar_txt_altura(caminho_data: str):
    # Busca recursiva por .txt em todo o diretório do DIA (similar ao encontrar_pasta_rinex)
    import os
    candidatos_nome = []
    candidatos_qualquer = []
    tokens = ("antena", "antea", "altura")

    for raiz, dirs, files in os.walk(caminho_data):
        for f in files:
            if f.lower().endswith(".txt"):
                caminho = os.path.join(raiz, f)
                nome_norm = _strip_accents(f).lower().replace(' ', '')
                if any(t in nome_norm for t in tokens):
                    candidatos_nome.append(caminho)
                else:
                    candidatos_qualquer.append(caminho)

    if candidatos_nome:
        return candidatos_nome
    return candidatos_qualquer

import os
import re
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook

def _set_merged_safe(ws, cell_address: str, value):
    """Set value on a cell that may be inside a merged range by writing to the range's top-left cell."""
    from openpyxl.utils import coordinate_to_tuple, get_column_letter
    r, c = coordinate_to_tuple(cell_address)
    for mr in ws.merged_cells.ranges:
        if (r, c) in mr:
            # Write to the top-left cell of the merged range
            top_left = ws.cell(row=mr.min_row, column=mr.min_col)
            top_left.value = value
            return
    # Not in a merged range -> write normally
    ws[cell_address] = value


def gerar_planilha_campo(
    dir_rinex_base=r"\\192.168.2.28\\i\\80225_PROJETO_IAT_PARANA\\4 Ponto de apoio\\RINEX",
    dir_template=r"\\192.168.2.28\i\80225_PROJETO_IAT_PARANA\3 Execução de voo",
    dir_saida=r"\\192.168.2.28\i\80225_PROJETO_IAT_PARANA\3 Execução de voo\ENTREGA\2_APOIO_DE_CAMPO",
    datas_filtradas=None,
    log_func=print
):
    processados = []
    datas_filtradas = set(datas_filtradas) if datas_filtradas else None
    registro_cache = None
    arquivo_registro = os.path.join(dir_saida, "dias_processados.txt")
    os.makedirs(dir_saida, exist_ok=True)
    MUNICIPIOS = {
        "153": "Clevelândia",
        "154": "Chopinzinho",
        "155": "Laranjeiras do Sul",
    }

    BASES = {
        "153": "BV-153",
        "154": "BV-154",
        "155": "BV-155",
    }

    def encontrar_pasta_rinex(caminho_data):
        for raiz, dirs, _ in os.walk(caminho_data):
            for dir_nome in dirs:
                if dir_nome.lower() == "rinex":
                    return os.path.join(raiz, dir_nome)
        return None

    def carregar_registro():
        if os.path.exists(arquivo_registro):
            registros = {}
            with open(arquivo_registro, 'r', encoding='utf-8') as f:
                for linha in f:
                    linha = linha.strip()
                    if not linha:
                        continue
                    partes = linha.split(';')
                    if len(partes) == 2:
                        registros[partes[0]] = partes[1]
            return registros
        return {}

    def atualizar_registro(data_dt, juliano):
        nonlocal registro_cache
        if registro_cache is None:
            registro_cache = carregar_registro()
        data_str = data_dt.strftime('%Y-%m-%d')
        registro_cache[data_str] = str(juliano)
        with open(arquivo_registro, 'w', encoding='utf-8') as f:
            for chave in sorted(registro_cache):
                f.write(f"{chave};{registro_cache[chave]}\n")

    def ajustar_hora(hora_str):
        h, m, s = map(int, hora_str.split(":"))
        nova = datetime(2000, 1, 1, h, m, s) - timedelta(hours=3)
        return nova.strftime("%H:%M:%S")

    def obter_base_id(nome_base):
        match = re.search(r"(\d{3})", nome_base)
        return match.group(1) if match else None

    for base_folder in sorted(os.listdir(dir_rinex_base)):
        caminho_base = os.path.join(dir_rinex_base, base_folder)
        if not os.path.isdir(caminho_base):
            continue

        base_id = obter_base_id(base_folder)
        if base_id not in BASES:
            log_func(f"[AVISO] Base não reconhecida: {base_folder}")
            continue

        base_nome = BASES[base_id]
        municipio = MUNICIPIOS.get(base_id, "Munic. Desconhecido")

        for data_folder in sorted(os.listdir(caminho_base)):
            if not data_folder.startswith("2025"):
                continue

            caminho_data = os.path.join(caminho_base, data_folder)
            if not os.path.isdir(caminho_data):
                continue

            try:
                data_dt = datetime.strptime(data_folder, "%Y%m%d")
            except ValueError:
                log_func(f"[AVISO] Nome de pasta de dia inválido em {base_nome}: {data_folder}")
                continue

            if datas_filtradas and data_dt.date() not in datas_filtradas:
                continue

            log_func(f"[INFO] Processando Campo Classe I — {base_nome} — {data_dt.strftime('%Y-%m-%d')}...")
            rinex_dir = encontrar_pasta_rinex(caminho_data)

            if not rinex_dir:
                log_func(f"[AVISO] Pasta 'Rinex' não encontrada em {caminho_data} (procura em subpastas)")
                continue

            arquivos = os.listdir(rinex_dir)
            arquivo_obs = next((f for f in arquivos if f.lower().endswith((".o", ".25o"))), None)
            if not arquivo_obs:
                log_func(f"[AVISO] Nenhum arquivo .o ou .25o encontrado em {rinex_dir}")
                continue

            caminho_obs = os.path.join(rinex_dir, arquivo_obs)
            with open(caminho_obs, "r") as f:
                linhas = f.readlines()

            hora_inicio = hora_fim = None
            for linha in linhas:
                if "TIME OF FIRST OBS" in linha:
                    p = linha.split()
                    hora_inicio = f"{int(p[3]):02d}:{int(p[4]):02d}:{int(float(p[5])):02d}"
                elif "TIME OF LAST OBS" in linha:
                    p = linha.split()
                    hora_fim = f"{int(p[3]):02d}:{int(p[4]):02d}:{int(float(p[5])):02d}"

            if not hora_inicio or not hora_fim:
                log_func("[AVISO] Horários não encontrados. Pulando.")
                continue

            hora_inicio = ajustar_hora(hora_inicio)
            hora_fim = ajustar_hora(hora_fim)

            juliano = data_dt.timetuple().tm_yday
            ano = data_dt.year

            nome_template = "GPS_0231_DIA_xxx.xlsx"
            caminho_template = os.path.join(dir_template, nome_template)
            if not os.path.exists(caminho_template):
                log_func(f"[ERRO] Template não encontrado: {nome_template}")
                continue

            wb = load_workbook(caminho_template)
            ws = wb.active
            ws["J8"] = data_dt.strftime("%d/%m/%Y")
            ws["K9"] = municipio
            ws["B15"] = base_nome
            ws["C11"] = f"{juliano}_{data_dt.year}"
            log_func(f"[OK] Dia juliano (C11) = {juliano}_{data_dt.year}")
            # === Altura da antena (TXT no dia) ===
            DEFAULT_ALTURA = None
            altura_txts = _procurar_txt_altura(caminho_data)
            altura_str = None
            altura_origem = None
            if altura_txts:
                # Tente cada candidato até achar um valor
                for _txt in sorted(altura_txts):
                    try:
                        with open(_txt, 'r', encoding='utf-8', errors='ignore') as fh:
                            conteudo = fh.read()
                    except Exception:
                        try:
                            with open(_txt, 'r', encoding='latin-1', errors='ignore') as fh:
                                conteudo = fh.read()
                        except Exception:
                            conteudo = ""
                    valor = _extrair_altura_do_txt(conteudo) if conteudo else None
                    if valor is not None:
                        altura_str = f"{valor:.3f} m".replace(".", ",")
                        altura_origem = _txt
                        break
            if altura_str:
                ws["D15"] = altura_str
                ws["E15"] = altura_str
                log_func(f"[OK] Altura da antena lida de '{os.path.basename(altura_origem)}': {altura_str}")
            else:
                ws["D15"] = ""
                ws["E15"] = ""
                log_func(f"[AVISO] Altura da antena não identificada em {caminho_data}. Campo deixado em branco.")
            ws["G15"] = hora_inicio
            ws["H15"] = hora_fim

            pasta_saida = os.path.join(dir_saida, f"DIA_{juliano}_{data_dt.year}", "GPS_0231")
            os.makedirs(pasta_saida, exist_ok=True)

            nome_saida = f"GPS_0231_DIA_{juliano}.xlsx"
            wb.save(os.path.join(pasta_saida, nome_saida))

            for f in arquivos:
                origem = os.path.join(rinex_dir, f)
                destino = os.path.join(pasta_saida, f)
                shutil.copy2(origem, destino)

            atualizar_registro(data_dt, juliano)
            log_func(f"[OK] Planilha de campo gerada para {base_nome} em {data_dt.strftime('%d/%m/%Y')}.")
            processados.append((data_dt.date(), base_nome, pasta_saida))

    return processados
