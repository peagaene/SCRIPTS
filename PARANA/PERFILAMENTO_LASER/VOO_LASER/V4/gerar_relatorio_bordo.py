import os
from collections import defaultdict
from datetime import datetime, timedelta

import geopandas as gpd
import pandas as pd
from openpyxl import load_workbook

EPSG_UTM22S = "EPSG:31982"
CAMINHO_TEMPLATE_RELATORIO = r"\\192.168.2.28\i\80225_PROJETO_IAT_PARANA\3 Execução de voo\ES_RB_LASER_LXX_Y_R0_ddmmaaaa.xlsx"
CAMINHO_PLANILHA_VOO = r"\\192.168.2.28\i\80225_PROJETO_IAT_PARANA\3 Execução de voo\VOO.xlsx"


def gps_time_to_datetime(week, seconds):
    gps_epoch = datetime(1980, 1, 6)
    return gps_epoch + timedelta(weeks=int(week), seconds=float(seconds))


def normalizar_piloto(nome):
    nome = nome.upper().strip()
    if nome in ["TOFOLI", "TOFOLLI", "TOFFOLI"]:
        return "Gustavo Tofoli"
    elif nome in ["ZUCH", "ZUCHI", "ZUC"]:
        return "Ivandro Zuchi"
    elif nome in ["DELANO"]:
        return "Delano Riker"
    elif nome in ["CASTELANI"]:
        return "Gustavo Castelani"
    print(f"[AVISO] Nome de piloto não reconhecido: {nome}")
    return nome


def normalizar_operador(nome, dia):
    nome = nome.upper().strip()
    if nome == "ÁLVARO A.":
        return "Álvaro Antônio"
    elif nome == "GILMAR R.":
        return "Gilmar Rocha"
    elif nome == "KEVIN":
        return "Kevin Bueno"
    elif nome == "MARCUS":
        return "Marcus Moura"
    return nome.title()


def carregar_plano_voo(dir_plano_voo, lote_id, cache):
    if lote_id not in cache:
        caminho = os.path.join(dir_plano_voo, f"ES_PV_LASER_L{lote_id}_FAIXAS_R0.shp")
        if os.path.exists(caminho):
            cache[lote_id] = gpd.read_file(caminho).to_crs(EPSG_UTM22S)
        else:
            print(f"[AVISO] Plano de voo não encontrado para o lote {lote_id}: {caminho}")
            cache[lote_id] = None
    return cache[lote_id]


def gerar_relatorio_bordo(dir_execucao_voo, blocos_info, datas_filtro):
    df_voo = pd.read_excel(CAMINHO_PLANILHA_VOO, header=3)
    df_voo.columns = df_voo.columns.str.strip()
    df_voo["DATA VOO"] = pd.to_datetime(df_voo["DATA VOO"], errors="coerce")

    dir_plano_voo = os.path.join(dir_execucao_voo, "PLANO_DE_VOO")
    plano_cache = {}

    for info in blocos_info:
        dir_bloco = info["dir_bloco"]
        bloco_nome = info["bloco_nome"]
        bloco_letra = bloco_nome.split("_")[-1]
        lote_id = info["lote_id"]
        dir_lote = info["dir_lote"]

        if not os.path.isdir(dir_bloco):
            print(f"[AVISO] Diretório do bloco não encontrado: {dir_bloco}")
            continue

        plano_voo = carregar_plano_voo(dir_plano_voo, lote_id, plano_cache)
        if plano_voo is None or plano_voo.empty:
            continue

        faixas_por_dia = defaultdict(list)

        for pasta in sorted(os.listdir(dir_bloco)):
            if not pasta.startswith("2025"):
                continue
            data_str, _ = pasta.split("_", 1)
            data_dt = datetime.strptime(data_str, "%Y%m%d")
            if datas_filtro and data_dt.date() not in datas_filtro:
                continue

            dir_dia = os.path.join(dir_bloco, pasta)
            shp_voo = [f for f in os.listdir(dir_dia) if f.endswith(".shp") and "_TRJ" not in f]
            if not shp_voo:
                print(f"[AVISO] Nenhum SHP de faixa encontrado em {dir_dia}")
                continue

            gdf_faixas = gpd.read_file(os.path.join(dir_dia, shp_voo[0])).to_crs(EPSG_UTM22S)

            for _, faixa in gdf_faixas.iterrows():
                centroide = faixa.geometry.centroid
                idx_min = plano_voo.distance(centroide).idxmin()
                linha_voo = plano_voo.iloc[idx_min]
                nome = linha_voo["FlightLine"]
                gps_week = faixa.get("GPS_week")
                gps_start = faixa.get("GPS_Start_")
                gps_stop = faixa.get("GPS_Stop_T")

                hora_inicio = (
                    gps_time_to_datetime(gps_week, gps_start).strftime("%Y-%m-%dT%H:%M:%S")
                    if pd.notna(gps_week) and pd.notna(gps_start)
                    else ""
                )
                hora_fim = (
                    gps_time_to_datetime(gps_week, gps_stop).strftime("%Y-%m-%dT%H:%M:%S")
                    if pd.notna(gps_week) and pd.notna(gps_stop)
                    else ""
                )

                altitude_m = faixa.get("Altitude_[")
                velocidade_nos = faixa.get("Speed_[m/s")

                altitude_pes = round(altitude_m * 3.28084, 1) if isinstance(altitude_m, (int, float)) else ""
                velocidade_nos = round(velocidade_nos * 1.94384, 1) if isinstance(velocidade_nos, (int, float)) else ""

                faixas_por_dia[data_dt].append(
                    [nome, hora_inicio, hora_fim, altitude_pes, velocidade_nos, 20, 80, 250]
                )

        for data_dt, dados_completos in faixas_por_dia.items():
            dados_completos.sort(key=lambda x: x[0])
            data_formatada = data_dt.strftime("%d%m%Y")
            dia_juliano = data_dt.timetuple().tm_yday

            dir_saida = os.path.join(dir_lote, "ENTREGA", "3_RELATORIO_BORDO", bloco_nome)
            os.makedirs(dir_saida, exist_ok=True)
            caminho_saida_excel = os.path.join(
                dir_saida, f"ES_RB_LASER_L{lote_id}_{bloco_letra}_R0_{data_formatada}.xlsx"
            )

            wb = load_workbook(CAMINHO_TEMPLATE_RELATORIO)
            ws = wb.active

            info_voo_dia = df_voo[df_voo["DATA VOO"] == data_dt]
            nome_piloto = nome_operador = obs_texto = ""
            if not info_voo_dia.empty:
                piloto_raw = str(info_voo_dia.iloc[0]["COMANDANTE"]).strip().upper()
                operador_raw = str(info_voo_dia.iloc[0]["OPERADOR"]).strip().upper()
                obs_raw = info_voo_dia.iloc[0].get("OBSERVAÇÕES", "")

                nome_piloto = f"NOME DO PILOTO: {normalizar_piloto(piloto_raw)}"
                nome_operador = f"NOME DO OPERADOR: {normalizar_operador(operador_raw, data_dt)}"
                obs_texto = obs_raw.capitalize() if isinstance(obs_raw, str) and obs_raw.strip() else ""

            ws["I13"] = nome_piloto
            ws["I14"] = nome_operador
            ws.cell(row=12, column=9).value = f"DIA JULIANO: {dia_juliano}"
            ws.cell(row=15, column=9).value = f"DATA: {data_dt.strftime('%d/%m/%Y')}"
            ws.cell(row=13, column=7).value = f"NUM. FAIXAS: {len(dados_completos)}"

            for i, linha in enumerate(dados_completos):
                linha_excel = 17 + i
                for j, valor in enumerate(linha):
                    ws.cell(row=linha_excel, column=2 + j).value = valor
                ws.cell(row=linha_excel, column=10).value = "boa"
                ws.cell(row=linha_excel, column=11).value = obs_texto

            wb.save(caminho_saida_excel)
            print(f"[OK] Planilha de Relatório de Bordo salva: {caminho_saida_excel}")
